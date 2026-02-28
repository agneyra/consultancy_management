from flask import Blueprint, render_template, request, send_file, redirect, url_for, flash, jsonify
from flask_login import login_required, current_user
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from utils.decorators import agent_required
from utils.excel_handler import (
    export_students_to_excel,
    import_students_from_excel,
    normalize_email,
    normalize_phone,
    is_valid_email
)
from models.database import db
from models.student import Student, StudentSemesterFee
from models.consultancy import Consultancy
from models.transaction import Transaction, Announcement
from models.user import AgentConsultancy, User
from sqlalchemy import func
from io import BytesIO
import os
import pandas as pd
import json
from datetime import datetime
from models.transaction import ChangeLog

agent_bp = Blueprint('agent', __name__)


def get_agent_consultancy_ids(agent_user):
    ids = set(agent_user.get_accessible_consultancy_ids())

    # Backward-compatible fallback for legacy records.
    if agent_user.consultancy_id:
        ids.add(agent_user.consultancy_id)

    # Ensure legacy primary mapping exists in link table.
    if agent_user.consultancy_id and not AgentConsultancy.query.filter_by(
        agent_id=agent_user.id,
        consultancy_id=agent_user.consultancy_id
    ).first():
        db.session.add(AgentConsultancy(agent_id=agent_user.id, consultancy_id=agent_user.consultancy_id))
        db.session.commit()

    return list(ids)


def validate_student_contact_uniqueness(email, phone, exclude_student_id=None):
    """Return an error message if student email/phone conflicts with existing records."""
    normalized_email = normalize_email(email)
    normalized_phone = normalize_phone(phone)

    if not normalized_email:
        return 'Email is required!'
    if not is_valid_email(normalized_email):
        return 'Enter a valid email address (example: name@example.com)!'
    if not normalized_phone:
        return 'Phone number is required!'

    for consultancy in Consultancy.query.all():
        if normalize_email(consultancy.email) == normalized_email:
            return f'Email already used by hostel "{consultancy.name}"!'
        if normalize_phone(consultancy.phone) == normalized_phone:
            return f'Phone number already used by hostel "{consultancy.name}"!'

    students_query = Student.query
    if exclude_student_id:
        students_query = students_query.filter(Student.id != exclude_student_id)
    for student in students_query.all():
        if normalize_email(student.email) == normalized_email:
            return 'Email already used by another student!'
        if normalize_phone(student.phone) == normalized_phone:
            return 'Phone number already used by another student!'

    excluded_user_id = None
    if exclude_student_id:
        excluded_student = Student.query.get(exclude_student_id)
        if excluded_student:
            excluded_user_id = excluded_student.user_id

    for user in User.query.all():
        if excluded_user_id and user.id == excluded_user_id:
            continue
        if normalize_email(user.email) == normalized_email:
            return 'Email already used by another account!'
        if normalize_phone(user.phone) == normalized_phone:
            return 'Phone number already used by another account!'

    return None


def parse_semester_value(raw_value):
    try:
        if raw_value is None:
            return None
        text_value = str(raw_value).strip()
        if not text_value:
            return None
        numeric_value = float(text_value)
        if not numeric_value.is_integer():
            return None
        semester = int(numeric_value)
    except (TypeError, ValueError):
        return None
    if 1 <= semester <= 8:
        return semester
    return None


def parse_prn_value(raw_value):
    if raw_value is None or pd.isna(raw_value):
        return ''
    if isinstance(raw_value, (int, float)):
        numeric_value = float(raw_value)
        if numeric_value.is_integer():
            return str(int(numeric_value))
    text_value = str(raw_value).strip()
    return '' if text_value.lower() == 'nan' else text_value


def read_uploaded_dataframe(filepath, filename):
    if filename.lower().endswith('.csv'):
        return pd.read_csv(filepath)

    try:
        return pd.read_excel(filepath)
    except Exception as read_error:
        # Pandas 3+ may reject older openpyxl versions. Fallback to direct openpyxl read.
        if 'openpyxl' not in str(read_error).lower():
            raise

        workbook = None
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(filepath, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return pd.DataFrame()

            headers = []
            for header in rows[0]:
                if header is None:
                    headers.append('')
                else:
                    headers.append(str(header).strip())

            data_rows = []
            for row_values in rows[1:]:
                if row_values is None:
                    continue
                if all(
                    cell is None or (isinstance(cell, str) and not cell.strip())
                    for cell in row_values
                ):
                    continue

                row_dict = {}
                for index, header in enumerate(headers):
                    if not header:
                        continue
                    row_dict[header] = row_values[index] if index < len(row_values) else None
                data_rows.append(row_dict)

            return pd.DataFrame(data_rows)
        except Exception:
            raise read_error
        finally:
            if workbook is not None:
                workbook.close()


def semester_fee_payload(semester_fee):
    return {
        'id': semester_fee.id,
        'student_id': semester_fee.student_id,
        'semester': semester_fee.semester,
        'total_fees': semester_fee.total_fees,
        'fees_paid': semester_fee.fees_paid,
        'fees_pending': semester_fee.fees_pending
    }


def log_change(user_id, user_role, action, table, record_id, changes):
    """Log changes to database"""
    try:
        log_entry = ChangeLog(
            user_id=user_id,
            user_role=user_role,
            action=action,
            table_name=table,
            record_id=record_id,
            changes=json.dumps(changes),
            timestamp=datetime.utcnow()
        )
        db.session.add(log_entry)
        db.session.commit()
    except Exception as e:
        print(f"Logging error: {str(e)}")
        pass  

@agent_bp.route('/dashboard')
@login_required
@agent_required
def dashboard():
    consultancy_ids = get_agent_consultancy_ids(current_user)
    consultancies = (
        Consultancy.query
        .filter(Consultancy.id.in_(consultancy_ids))
        .order_by(Consultancy.hostel_code.asc())
        .all()
        if consultancy_ids else []
    )

    selected_hostel_id = request.args.get('hostel_id', type=int)
    if selected_hostel_id not in consultancy_ids:
        selected_hostel_id = None

    filtered_consultancy_ids = [selected_hostel_id] if selected_hostel_id else consultancy_ids
    selected_semester = request.args.get('semester', type=int)
    if selected_semester not in range(1, 9):
        selected_semester = None

    students = (
        Student.query.filter(Student.consultancy_id.in_(filtered_consultancy_ids)).all()
        if filtered_consultancy_ids else []
    )
    total_students = len(students)

    if selected_semester:
        semester_rows = (
            StudentSemesterFee.query
            .join(Student, Student.id == StudentSemesterFee.student_id)
            .filter(
                Student.consultancy_id.in_(filtered_consultancy_ids),
                StudentSemesterFee.semester == selected_semester
            )
            .all()
        ) if filtered_consultancy_ids else []
        total_fees = sum(row.total_fees for row in semester_rows)
        fees_paid = sum(row.fees_paid for row in semester_rows)
        fees_pending = total_fees - fees_paid
        total_students = len({row.student_id for row in semester_rows})
    else:
        total_fees = sum(s.total_fees for s in students)
        fees_paid = sum(s.fees_paid for s in students)
        fees_pending = total_fees - fees_paid
    
    # Get active announcements
    announcements = Announcement.query.filter_by(is_active=True).order_by(Announcement.created_at.desc()).all()
    
    stats = {
        'total_fees': total_fees,
        'fees_paid': fees_paid,
        'fees_pending': fees_pending,
        'total_students': total_students
    }

    visible_consultancies = (
        [c for c in consultancies if c.id == selected_hostel_id]
        if selected_hostel_id else consultancies
    )
    hostel_blocks = []
    for consultancy in visible_consultancies:
        hostel_students = [s for s in students if s.consultancy_id == consultancy.id]
        if selected_semester:
            hostel_semester_rows = (
                StudentSemesterFee.query
                .join(Student, Student.id == StudentSemesterFee.student_id)
                .filter(
                    Student.consultancy_id == consultancy.id,
                    StudentSemesterFee.semester == selected_semester
                )
                .all()
            )
            hostel_total = sum(row.total_fees for row in hostel_semester_rows)
            hostel_paid = sum(row.fees_paid for row in hostel_semester_rows)
            hostel_student_count = len({row.student_id for row in hostel_semester_rows})
        else:
            hostel_total = sum(s.total_fees for s in hostel_students)
            hostel_paid = sum(s.fees_paid for s in hostel_students)
            hostel_student_count = len(hostel_students)
        hostel_blocks.append({
            'hostel_code': consultancy.hostel_code,
            'hostel_name': consultancy.name,
            'students': hostel_student_count,
            'total_fees': hostel_total,
            'fees_paid': hostel_paid,
            'fees_pending': hostel_total - hostel_paid
        })

    return render_template(
        'agent/dashboard.html',
        stats=stats,
        announcements=announcements,
        hostel_blocks=hostel_blocks,
        selected_semester=selected_semester,
        selected_hostel_id=selected_hostel_id,
        consultancies=consultancies
    )

@agent_bp.route('/students')
@login_required
@agent_required
def students_data():
    consultancy_ids = get_agent_consultancy_ids(current_user)
    consultancies = Consultancy.query.filter(Consultancy.id.in_(consultancy_ids)).order_by(Consultancy.hostel_code.asc()).all() if consultancy_ids else []
    selected_hostel_id = request.args.get('hostel_id', type=int)
    pending_filter = request.args.get('pending_filter', '')
    search = request.args.get('search', '')
    
    # Base query
    query = Student.query.filter(Student.consultancy_id.in_(consultancy_ids)) if consultancy_ids else Student.query.filter(Student.id == -1)

    if selected_hostel_id and selected_hostel_id in consultancy_ids:
        query = query.filter(Student.consultancy_id == selected_hostel_id)
    
    # Apply pending fee filter
    if pending_filter == 'has_pending':
        query = query.filter(Student.total_fees > Student.fees_paid)
    elif pending_filter == 'no_pending':
        query = query.filter(Student.total_fees <= Student.fees_paid)
    
    # Apply search filter
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            db.or_(
                Student.prn.like(search_term),
                Student.full_name.like(search_term),
                Student.email.like(search_term),
                Student.branch.like(search_term)
            )
        )
    
    students = query.all()
    
    return render_template(
        'agent/students_data.html',
        students=students,
        consultancies=consultancies,
        selected_hostel_id=selected_hostel_id
    )


@agent_bp.route('/students/add')
@login_required
@agent_required
def add_student_page():
    consultancy_ids = get_agent_consultancy_ids(current_user)
    consultancies = (
        Consultancy.query
        .filter(Consultancy.id.in_(consultancy_ids), Consultancy.is_active.is_(True))
        .order_by(Consultancy.hostel_code.asc())
        .all()
        if consultancy_ids else []
    )
    return render_template('agent/add_student.html', consultancies=consultancies)


@agent_bp.route('/students/sample-template')
@login_required
@agent_required
def download_sample_template():
    consultancy_ids = get_agent_consultancy_ids(current_user)
    consultancies = (
        Consultancy.query
        .filter(Consultancy.id.in_(consultancy_ids), Consultancy.is_active.is_(True))
        .order_by(Consultancy.hostel_code.asc())
        .all()
        if consultancy_ids else []
    )
    default_hostel_code = consultancies[0].hostel_code if consultancies else 'B1'

    df = pd.DataFrame([
        {
            'PRN': '2023001',
            'Name': 'John Doe',
            'Branch': 'Computer Science',
            'Email': 'john@example.com',
            'Phone': '9876543210',
            'Hostel_Code': default_hostel_code,
            'Semester': 1,
            'Total_Fees': 50000,
            'Fees_Paid': 10000
        }
    ])

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sample')
    output.seek(0)

    return send_file(
        output,
        download_name='students_sample_template.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@agent_bp.route('/students/semester-fees/sample-template')
@login_required
@agent_required
def download_semester_fee_template():
    df = pd.DataFrame([
        {
            'PRN': '2023001',
            'Semester': 1,
            'Total_Fees': 50000,
            'Fees_Paid': 10000
        }
    ])

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sample')
    output.seek(0)

    return send_file(
        output,
        download_name='semester_fee_update_template.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@agent_bp.route('/students/semester-fees/upload', methods=['POST'])
@login_required
@agent_required
def upload_semester_fees():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file uploaded'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No file selected'}), 400

    consultancy_ids = get_agent_consultancy_ids(current_user)
    if not consultancy_ids:
        return jsonify({
            'success': False,
            'message': 'No hostel is assigned to your account.'
        }), 400

    filename = secure_filename(file.filename)
    if not filename:
        return jsonify({'success': False, 'message': 'Invalid file name'}), 400
    upload_dir = os.path.join('static', 'uploads')
    os.makedirs(upload_dir, exist_ok=True)
    filepath = os.path.join(upload_dir, filename)

    results = {
        'processed': 0,
        'added': 0,
        'updated': 0,
        'failed': 0,
        'errors': []
    }

    try:
        file.save(filepath)

        df = read_uploaded_dataframe(filepath, filename)

        df.columns = [str(col).strip() for col in df.columns]
        column_aliases = {
            'Total_Fee': 'Total_Fees',
            'Fee_Paid': 'Fees_Paid',
            'Total Fee': 'Total_Fees',
            'Fee Paid': 'Fees_Paid'
        }
        for source_col, target_col in column_aliases.items():
            if source_col in df.columns and target_col not in df.columns:
                df.rename(columns={source_col: target_col}, inplace=True)

        required_columns = ['PRN', 'Semester', 'Total_Fees', 'Fees_Paid']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return jsonify({
                'success': False,
                'message': f"Missing required column(s): {', '.join(missing_columns)}"
            }), 400

        for index, row in df.iterrows():
            row_number = index + 2
            results['processed'] += 1
            try:
                prn = parse_prn_value(row.get('PRN'))
                if not prn:
                    raise ValueError('PRN is required')

                semester = parse_semester_value(row.get('Semester'))
                if semester is None:
                    raise ValueError('Semester must be between 1 and 8')

                total_fees_raw = row.get('Total_Fees', 0)
                fees_paid_raw = row.get('Fees_Paid', 0)

                total_fees = 0.0 if pd.isna(total_fees_raw) else float(total_fees_raw)
                fees_paid = 0.0 if pd.isna(fees_paid_raw) else float(fees_paid_raw)

                if total_fees < 0 or fees_paid < 0:
                    raise ValueError('Fees cannot be negative')
                if fees_paid > total_fees:
                    raise ValueError('Fees paid cannot be greater than total fees')

                student = (
                    Student.query
                    .filter(
                        Student.prn == prn,
                        Student.consultancy_id.in_(consultancy_ids)
                    )
                    .first()
                )
                if not student:
                    raise ValueError(f'Student with PRN {prn} not found in your assigned hostels')

                semester_fee = StudentSemesterFee.query.filter_by(
                    student_id=student.id,
                    semester=semester
                ).first()

                if semester_fee:
                    semester_fee.total_fees = total_fees
                    semester_fee.fees_paid = fees_paid
                    results['updated'] += 1
                else:
                    db.session.add(StudentSemesterFee(
                        student_id=student.id,
                        semester=semester,
                        total_fees=total_fees,
                        fees_paid=fees_paid
                    ))
                    results['added'] += 1

                student.recalculate_fee_totals()
                db.session.commit()
            except Exception as row_error:
                db.session.rollback()
                results['failed'] += 1
                results['errors'].append(f'Row {row_number}: {str(row_error)}')

        return jsonify({
            'success': True,
            'message': (
                'Semester fee upload complete! '
                f"Added: {results['added']}, Updated: {results['updated']}, Failed: {results['failed']}"
            ),
            'details': results
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error processing file: {str(e)}'}), 400
    finally:
        if os.path.exists(filepath):
            try:
                os.remove(filepath)
            except OSError:
                pass


@agent_bp.route('/students/upload', methods=['POST'])
@login_required
@agent_required
def upload_students():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file uploaded'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No file selected'}), 400

    consultancy_ids = get_agent_consultancy_ids(current_user)
    consultancies = (
        Consultancy.query
        .filter(Consultancy.id.in_(consultancy_ids), Consultancy.is_active.is_(True))
        .all()
        if consultancy_ids else []
    )
    allowed_hostel_codes = [c.hostel_code for c in consultancies]
    if not allowed_hostel_codes:
        return jsonify({
            'success': False,
            'message': 'No active hostel is assigned to your account.'
        }), 400

    filename = secure_filename(file.filename)
    if not filename:
        return jsonify({'success': False, 'message': 'Invalid file name'}), 400

    upload_dir = os.path.join('static', 'uploads')
    os.makedirs(upload_dir, exist_ok=True)
    filepath = os.path.join(upload_dir, filename)

    try:
        file.save(filepath)
        success, result = import_students_from_excel(
            filepath,
            allowed_hostel_codes=allowed_hostel_codes
        )

        if success:
            return jsonify({
                'success': True,
                'message': f"Import complete! Success: {result['success']}, Failed: {result['failed']}",
                'details': result
            })
        return jsonify({'success': False, 'message': result}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error processing file: {str(e)}'}), 400
    finally:
        if os.path.exists(filepath):
            try:
                os.remove(filepath)
            except OSError:
                pass


@agent_bp.route('/students/add-single', methods=['POST'])
@login_required
@agent_required
def add_single_student():
    try:
        prn = (request.form.get('prn') or '').strip()
        full_name = (request.form.get('full_name') or '').strip()
        branch = (request.form.get('branch') or '').strip()
        email = normalize_email(request.form.get('email'))
        phone = normalize_phone(request.form.get('phone'))
        consultancy_id_raw = request.form.get('consultancy_id')
        semester_raw = request.form.get('semester')
        total_fees = float(request.form.get('total_fees') or 0)
        fees_paid = float(request.form.get('fees_paid') or 0)

        if not prn or not full_name or not branch:
            flash('PRN, full name, and branch are required!', 'error')
            return redirect(url_for('agent.add_student_page'))

        try:
            consultancy_id = int(consultancy_id_raw)
        except (TypeError, ValueError):
            flash('Please select a valid hostel.', 'error')
            return redirect(url_for('agent.add_student_page'))

        semester = parse_semester_value(semester_raw)
        if semester is None:
            flash('Please select a valid semester (1 to 8).', 'error')
            return redirect(url_for('agent.add_student_page'))

        consultancy_ids = set(get_agent_consultancy_ids(current_user))
        if consultancy_id not in consultancy_ids:
            flash('You can add students only to your assigned hostels.', 'error')
            return redirect(url_for('agent.add_student_page'))

        consultancy = Consultancy.query.get(consultancy_id)
        if not consultancy or not consultancy.is_active:
            flash('Selected hostel is unavailable or deactivated.', 'error')
            return redirect(url_for('agent.add_student_page'))

        existing_student = Student.query.filter_by(prn=prn).first()
        if existing_student:
            flash('Student with this PRN already exists!', 'error')
            return redirect(url_for('agent.add_student_page'))

        contact_error = validate_student_contact_uniqueness(email=email, phone=phone)
        if contact_error:
            flash(contact_error, 'error')
            return redirect(url_for('agent.add_student_page'))

        if fees_paid > total_fees:
            flash('Fees paid cannot be greater than total fees.', 'error')
            return redirect(url_for('agent.add_student_page'))

        user = User(
            username=prn,
            password=generate_password_hash(phone),
            email=email,
            phone=phone,
            role='student',
            consultancy_id=consultancy_id
        )
        db.session.add(user)
        db.session.flush()

        student = Student(
            user_id=user.id,
            consultancy_id=consultancy_id,
            prn=prn,
            full_name=full_name,
            branch=branch,
            email=email,
            phone=phone,
            total_fees=0,
            fees_paid=0
        )
        db.session.add(student)
        db.session.flush()

        student.semester_fees.append(StudentSemesterFee(
            semester=semester,
            total_fees=total_fees,
            fees_paid=fees_paid
        ))
        student.recalculate_fee_totals()
        db.session.commit()

        flash(
            f'Student {full_name} added successfully! Login: Username={prn}, Password={phone}',
            'success'
        )
        return redirect(url_for('agent.add_student_page'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding student: {str(e)}', 'error')
        return redirect(url_for('agent.add_student_page'))


@agent_bp.route('/students/<int:student_id>/semester-fees/add', methods=['POST'])
@login_required
@agent_required
def add_semester_fee(student_id):
    consultancy_ids = get_agent_consultancy_ids(current_user)
    student = Student.query.filter(
        Student.id == student_id,
        Student.consultancy_id.in_(consultancy_ids)
    ).first_or_404()

    data = request.get_json() or {}
    semester = parse_semester_value(data.get('semester'))
    if semester is None:
        return jsonify({'success': False, 'message': 'Semester must be between 1 and 8'}), 400

    try:
        total_fees = float(data.get('total_fees', 0))
        fees_paid = float(data.get('fees_paid', 0))
    except (TypeError, ValueError):
        return jsonify({'success': False, 'message': 'Fees must be valid numbers'}), 400

    if total_fees < 0 or fees_paid < 0:
        return jsonify({'success': False, 'message': 'Fees cannot be negative'}), 400
    if fees_paid > total_fees:
        return jsonify({'success': False, 'message': 'Paid fees cannot be greater than total fees'}), 400

    existing_fee = StudentSemesterFee.query.filter_by(student_id=student.id, semester=semester).first()
    if existing_fee:
        return jsonify({'success': False, 'message': 'This semester already exists for the student'}), 400

    try:
        semester_fee = StudentSemesterFee(
            student_id=student.id,
            semester=semester,
            total_fees=total_fees,
            fees_paid=fees_paid
        )
        db.session.add(semester_fee)
        db.session.flush()

        student.recalculate_fee_totals()
        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Semester fee added successfully',
            'semester_fee': semester_fee_payload(semester_fee),
            'student_totals': {
                'total_fees': student.total_fees,
                'fees_paid': student.fees_paid,
                'fees_pending': student.fees_pending
            }
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400


@agent_bp.route('/students/semester-fees/update/<int:semester_fee_id>', methods=['POST'])
@login_required
@agent_required
def update_semester_fee(semester_fee_id):
    consultancy_ids = get_agent_consultancy_ids(current_user)
    semester_fee = (
        StudentSemesterFee.query
        .join(Student, Student.id == StudentSemesterFee.student_id)
        .filter(
            StudentSemesterFee.id == semester_fee_id,
            Student.consultancy_id.in_(consultancy_ids)
        )
        .first_or_404()
    )

    student = semester_fee.student
    data = request.get_json() or {}

    semester = parse_semester_value(data.get('semester', semester_fee.semester))
    if semester is None:
        return jsonify({'success': False, 'message': 'Semester must be between 1 and 8'}), 400

    try:
        total_fees = float(data.get('total_fees', semester_fee.total_fees))
        fees_paid = float(data.get('fees_paid', semester_fee.fees_paid))
    except (TypeError, ValueError):
        return jsonify({'success': False, 'message': 'Fees must be valid numbers'}), 400

    if total_fees < 0 or fees_paid < 0:
        return jsonify({'success': False, 'message': 'Fees cannot be negative'}), 400
    if fees_paid > total_fees:
        return jsonify({'success': False, 'message': 'Paid fees cannot be greater than total fees'}), 400

    duplicate_semester = StudentSemesterFee.query.filter(
        StudentSemesterFee.student_id == student.id,
        StudentSemesterFee.semester == semester,
        StudentSemesterFee.id != semester_fee.id
    ).first()
    if duplicate_semester:
        return jsonify({'success': False, 'message': 'Another fee row already exists for this semester'}), 400

    try:
        semester_fee.semester = semester
        semester_fee.total_fees = total_fees
        semester_fee.fees_paid = fees_paid

        student.recalculate_fee_totals()
        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Semester fee updated successfully',
            'semester_fee': semester_fee_payload(semester_fee),
            'student_totals': {
                'total_fees': student.total_fees,
                'fees_paid': student.fees_paid,
                'fees_pending': student.fees_pending
            }
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400


@agent_bp.route('/students/semester-fees/delete/<int:semester_fee_id>', methods=['POST'])
@login_required
@agent_required
def delete_semester_fee(semester_fee_id):
    consultancy_ids = get_agent_consultancy_ids(current_user)
    semester_fee = (
        StudentSemesterFee.query
        .join(Student, Student.id == StudentSemesterFee.student_id)
        .filter(
            StudentSemesterFee.id == semester_fee_id,
            Student.consultancy_id.in_(consultancy_ids)
        )
        .first_or_404()
    )
    student = semester_fee.student

    try:
        db.session.delete(semester_fee)
        db.session.flush()
        student.recalculate_fee_totals()
        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Semester fee deleted successfully',
            'student_id': student.id,
            'student_totals': {
                'total_fees': student.total_fees,
                'fees_paid': student.fees_paid,
                'fees_pending': student.fees_pending
            }
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400


@agent_bp.route('/change-password', methods=['GET', 'POST'])
@login_required
@agent_required
def change_password():
    if request.method == 'POST':
        old_password = request.form.get('old_password', '')
        new_password = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')

        if not check_password_hash(current_user.password, old_password):
            flash('Old password is incorrect', 'error')
            return redirect(url_for('agent.change_password'))

        if len(new_password) < 6:
            flash('New password must be at least 6 characters', 'error')
            return redirect(url_for('agent.change_password'))

        if new_password != confirm_password:
            flash('New passwords do not match', 'error')
            return redirect(url_for('agent.change_password'))

        if old_password == new_password:
            flash('New password must be different from old password', 'error')
            return redirect(url_for('agent.change_password'))

        current_user.password = generate_password_hash(new_password)
        db.session.commit()
        flash('Password changed successfully!', 'success')
        return redirect(url_for('agent.dashboard'))

    return render_template('agent/change_password.html')

@agent_bp.route('/students/export')
@login_required
@agent_required
def export_students():
    consultancy_ids = get_agent_consultancy_ids(current_user)
    selected_hostel_id = request.args.get('hostel_id', type=int)
    query = Student.query.filter(Student.consultancy_id.in_(consultancy_ids)) if consultancy_ids else Student.query.filter(Student.id == -1)
    if selected_hostel_id and selected_hostel_id in consultancy_ids:
        query = query.filter(Student.consultancy_id == selected_hostel_id)
    students = query.all()
    
    df = export_students_to_excel(students)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Students')
    output.seek(0)
    
    return send_file(output,
                    download_name='students_data.xlsx',
                    as_attachment=True,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@agent_bp.route('/payment-history')
@login_required
@agent_required
def payment_history():
    flash('Payment History is disabled in the agent panel.', 'error')
    return redirect(url_for('agent.dashboard'))

@agent_bp.route('/payment-history/export')
@login_required
@agent_required
def export_payment_history():
    flash('Payment History export is disabled in the agent panel.', 'error')
    return redirect(url_for('agent.dashboard'))

@agent_bp.route('/students/update/<int:id>', methods=['POST'])
@login_required
@agent_required
def update_student(id):
    consultancy_ids = get_agent_consultancy_ids(current_user)
    student = Student.query.filter(Student.id == id, Student.consultancy_id.in_(consultancy_ids)).first_or_404()
    
    try:
        data = request.get_json() or {}

        prn = (data.get('prn') or student.prn).strip()
        full_name = (data.get('full_name') or student.full_name).strip()
        branch = (data.get('branch') or student.branch).strip()
        email = normalize_email(data.get('email', student.email))
        phone = normalize_phone(data.get('phone', student.phone))

        try:
            total_fees = float(data.get('total_fees', student.total_fees))
            fees_paid = float(data.get('fees_paid', student.fees_paid))
        except (TypeError, ValueError):
            return jsonify({'success': False, 'message': 'Total fees and paid fees must be valid numbers'}), 400

        try:
            consultancy_id = int(data.get('consultancy_id', student.consultancy_id))
        except (TypeError, ValueError):
            return jsonify({'success': False, 'message': 'Please select a valid hostel'}), 400

        if consultancy_id not in consultancy_ids:
            return jsonify({'success': False, 'message': 'You can only assign students to your own hostels'}), 403

        target_consultancy = Consultancy.query.filter(
            Consultancy.id == consultancy_id,
            Consultancy.is_active.is_(True)
        ).first()
        if not target_consultancy:
            return jsonify({'success': False, 'message': 'Selected hostel is not active'}), 400

        if fees_paid > total_fees:
            return jsonify({'success': False, 'message': 'Paid fees cannot be greater than total fees'}), 400

        duplicate_prn = Student.query.filter(Student.prn == prn, Student.id != student.id).first()
        if duplicate_prn:
            return jsonify({'success': False, 'message': 'Student with this PRN already exists'}), 400

        contact_error = validate_student_contact_uniqueness(
            email=email,
            phone=phone,
            exclude_student_id=student.id
        )
        if contact_error:
            return jsonify({'success': False, 'message': contact_error}), 400

        old_phone = student.phone

        # Log the change
        log_change(
            user_id=current_user.id,
            user_role='agent',
            action='update',
            table='students',
            record_id=student.id,
            changes=data
        )

        # Update student and linked user
        student.prn = prn
        student.full_name = full_name
        student.branch = branch
        student.email = email
        student.phone = phone
        student.total_fees = total_fees
        student.fees_paid = fees_paid
        student.consultancy_id = consultancy_id

        student.user.username = prn
        student.user.email = email
        student.user.phone = phone
        student.user.consultancy_id = consultancy_id
        if old_phone != phone:
            student.user.password = generate_password_hash(phone)

        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Student updated successfully',
            'student': {
                'id': student.id,
                'prn': student.prn,
                'full_name': student.full_name,
                'branch': student.branch,
                'email': student.email,
                'phone': student.phone,
                'consultancy_id': student.consultancy_id,
                'hostel_code': student.consultancy.hostel_code,
                'hostel_name': student.consultancy.name,
                'total_fees': student.total_fees,
                'fees_paid': student.fees_paid,
                'fees_pending': student.fees_pending
            }
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400

@agent_bp.route('/students/delete/<int:id>', methods=['POST'])
@login_required
@agent_required
def delete_student(id):
    consultancy_ids = get_agent_consultancy_ids(current_user)
    student = Student.query.filter(Student.id == id, Student.consultancy_id.in_(consultancy_ids)).first_or_404()
    user = student.user
    
    try:
        # Log the deletion
        log_change(
            user_id=current_user.id,
            user_role='agent',
            action='delete',
            table='students',
            record_id=student.id,
            changes={'student_name': student.full_name, 'prn': student.prn}
        )
        
        db.session.delete(student)
        db.session.delete(user)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Student deleted successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400

