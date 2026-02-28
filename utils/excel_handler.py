import pandas as pd
from werkzeug.security import generate_password_hash
from models.database import db
from models.user import User
from models.student import Student, StudentSemesterFee
from models.consultancy import Consultancy
import random
import string
import re
from decimal import Decimal, InvalidOperation
from utils.hostels import HOSTELS


def normalize_email(value):
    return ''.join(str(value or '').strip().lower().split())


def _coerce_integral_string(value):
    if value is None:
        return ''

    try:
        if pd.isna(value):
            return ''
    except TypeError:
        pass

    if isinstance(value, bool):
        return ''

    if isinstance(value, int):
        return str(value) if value >= 0 else ''

    if isinstance(value, float):
        if value < 0:
            return ''
        if value.is_integer():
            return str(int(value))
        return ''

    text = str(value).strip()
    if not text or text.lower() == 'nan':
        return ''

    text = text.replace(',', '')
    try:
        numeric_value = Decimal(text)
    except InvalidOperation:
        return ''

    if not numeric_value.is_finite() or numeric_value < 0:
        return ''
    if numeric_value != numeric_value.to_integral_value():
        return ''

    integral_value = numeric_value.to_integral_value()
    integral_text = format(integral_value, 'f')
    if integral_text.endswith('.0'):
        integral_text = integral_text[:-2]
    return integral_text


def normalize_phone(value):
    integral_text = _coerce_integral_string(value)
    if integral_text:
        return integral_text

    text = str(value or '').strip()
    if not text or text.lower() == 'nan':
        return ''
    return ''.join(ch for ch in text if ch.isdigit())


def normalize_prn(value):
    if value is None:
        return ''

    try:
        if pd.isna(value):
            return ''
    except TypeError:
        pass

    if isinstance(value, str):
        text_value = value.strip()
        if not text_value or text_value.lower() == 'nan':
            return ''
        if text_value.isdigit():
            return text_value
        integral_text = _coerce_integral_string(text_value)
        return integral_text or text_value

    integral_text = _coerce_integral_string(value)
    if integral_text:
        return integral_text

    text_value = str(value).strip()
    return '' if text_value.lower() == 'nan' else text_value


def is_valid_email(value):
    pattern = r'^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$'
    return bool(re.match(pattern, value or ''))


def generate_password(length=8):
    """Generate a random password"""
    characters = string.ascii_letters + string.digits
    return ''.join(random.choice(characters) for i in range(length))


def read_uploaded_dataframe(file_path):
    try:
        return pd.read_excel(file_path)
    except Exception as read_error:
        if 'openpyxl' not in str(read_error).lower():
            raise

        workbook = None
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(file_path, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return pd.DataFrame()

            headers = []
            for header in rows[0]:
                if header is None:
                    headers.append('')
                else:
                    headers.append(str(header).strip())

            data_rows = []
            for row_values in rows[1:]:
                if row_values is None:
                    continue
                if all(
                    cell is None or (isinstance(cell, str) and not cell.strip())
                    for cell in row_values
                ):
                    continue

                row_dict = {}
                for index, header in enumerate(headers):
                    if not header:
                        continue
                    row_dict[header] = row_values[index] if index < len(row_values) else None
                data_rows.append(row_dict)

            return pd.DataFrame(data_rows)
        except Exception:
            raise read_error
        finally:
            if workbook is not None:
                workbook.close()

# Updated import_students_from_excel function in utils/excel_handler.py

def import_students_from_excel(file_path, allowed_hostel_codes=None):
    """
    Import students from Excel file
    Expected columns: PRN, Name, Branch, Email, Phone, Hostel_Code, Semester, Total_Fees, Fees_Paid
    allowed_hostel_codes (optional): restrict import to these hostel codes.
    """
    try:
        allowed_hostel_codes = {
            str(code).strip().upper()
            for code in (allowed_hostel_codes or [])
            if str(code).strip()
        }

        df = read_uploaded_dataframe(file_path)
        
        # Strip whitespace from column names
        df.columns = [str(col).strip() for col in df.columns]

        column_aliases = {
            'Total_Fee': 'Total_Fees',
            'Total Fee': 'Total_Fees',
            'Fees Paid': 'Fees_Paid',
            'Fee_Paid': 'Fees_Paid',
            'Hostel Code': 'Hostel_Code'
        }
        for source_col, target_col in column_aliases.items():
            if source_col in df.columns and target_col not in df.columns:
                df.rename(columns={source_col: target_col}, inplace=True)
        
        required_columns = ['PRN', 'Name', 'Branch', 'Email', 'Phone', 'Hostel_Code', 'Semester', 'Total_Fees']
        
        # Check if all required columns exist
        for col in required_columns:
            if col not in df.columns:
                return False, f"Missing required column: {col}"
        
        results = {
            'success': 0,
            'failed': 0,
            'errors': [],
            'credentials': []
        }

        seen_emails = set()
        seen_phones = set()
        
        for index, row in df.iterrows():
            try:
                prn = normalize_prn(row.get('PRN'))
                if not prn:
                    results['failed'] += 1
                    results['errors'].append(f"Row {index+2}: PRN is required")
                    continue

                # Get or create consultancy
                hostel_code = str(row['Hostel_Code']).strip().upper()

                if hostel_code not in HOSTELS:
                    results['failed'] += 1
                    results['errors'].append(
                        f"Row {index+2}: Invalid hostel code '{hostel_code}'"
                    )
                    continue

                if allowed_hostel_codes and hostel_code not in allowed_hostel_codes:
                    results['failed'] += 1
                    results['errors'].append(
                        f"Row {index+2}: Hostel code '{hostel_code}' is not assigned to your account"
                    )
                    continue

                consultancy = Consultancy.query.filter_by(hostel_code=hostel_code).first()

                if not consultancy:
                    if allowed_hostel_codes:
                        results['failed'] += 1
                        results['errors'].append(
                            f"Row {index+2}: Hostel code '{hostel_code}' is not available in system"
                        )
                        continue
                    consultancy = Consultancy(
                        hostel_code=hostel_code,
                        name=HOSTELS[hostel_code],  # AUTO name from mapping
                        contact_person="Auto Imported",
                        email=f"{hostel_code.lower()}@auto.local",
                        phone="0000000000",
                        address="Auto created from Excel import",
                        is_active=True
                    )
                    db.session.add(consultancy)
                    db.session.flush()  # REQUIRED to get consultancy.id

                try:
                    semester = int(float(row.get('Semester')))
                except (TypeError, ValueError):
                    results['failed'] += 1
                    results['errors'].append(f"Row {index+2}: Semester must be a number from 1 to 8")
                    continue

                if semester < 1 or semester > 8:
                    results['failed'] += 1
                    results['errors'].append(f"Row {index+2}: Semester must be between 1 and 8")
                    continue

                # Check if student already exists
                existing_student = Student.query.filter_by(prn=prn).first()
                if existing_student:
                    results['failed'] += 1
                    results['errors'].append(f"Row {index+2}: Student with PRN {prn} already exists")
                    continue

                # Get phone number from Excel
                phone_number = normalize_phone(row.get('Phone', ''))
                student_email = normalize_email(row['Email'])
                student_name = str(row.get('Name', '')).strip()
                student_branch = str(row.get('Branch', '')).strip()

                if not student_name or student_name.lower() == 'nan':
                    results['failed'] += 1
                    results['errors'].append(f"Row {index+2}: Name is required")
                    continue

                if not student_branch or student_branch.lower() == 'nan':
                    results['failed'] += 1
                    results['errors'].append(f"Row {index+2}: Branch is required")
                    continue
                
                # Validate phone number exists
                if not phone_number:
                    results['failed'] += 1
                    results['errors'].append(f"Row {index+2}: Phone number is required")
                    continue

                if not student_email:
                    results['failed'] += 1
                    results['errors'].append(f"Row {index+2}: Email is required")
                    continue
                if not is_valid_email(student_email):
                    results['failed'] += 1
                    results['errors'].append(
                        f"Row {index+2}: Invalid email format '{student_email}'. Use format like name@example.com"
                    )
                    continue

                # Block duplicate contact values inside the same uploaded file.
                if student_email in seen_emails:
                    results['failed'] += 1
                    results['errors'].append(f"Row {index+2}: Email {student_email} is duplicated in this file")
                    continue
                if phone_number in seen_phones:
                    results['failed'] += 1
                    results['errors'].append(f"Row {index+2}: Phone {phone_number} is duplicated in this file")
                    continue

                row_error = None

                # Block if email/phone already used by any hostel.
                for existing_consultancy in Consultancy.query.all():
                    if normalize_email(existing_consultancy.email) == student_email:
                        row_error = f"Row {index+2}: Email {student_email} is already used by hostel {existing_consultancy.name}"
                        break
                    if normalize_phone(existing_consultancy.phone) == phone_number:
                        row_error = f"Row {index+2}: Phone {phone_number} is already used by hostel {existing_consultancy.name}"
                        break

                # Block if email/phone already used by any student.
                if not row_error:
                    for existing_student in Student.query.all():
                        if normalize_email(existing_student.email) == student_email:
                            row_error = f"Row {index+2}: Email {student_email} is already used by another student"
                            break
                        if normalize_phone(existing_student.phone) == phone_number:
                            row_error = f"Row {index+2}: Phone {phone_number} is already used by another student"
                            break

                # Block if email/phone already used by any account.
                if not row_error:
                    for existing_user in User.query.all():
                        if normalize_email(existing_user.email) == student_email:
                            row_error = f"Row {index+2}: Email {student_email} is already used by another account"
                            break
                        if normalize_phone(existing_user.phone) == phone_number:
                            row_error = f"Row {index+2}: Phone {phone_number} is already used by another account"
                            break

                if row_error:
                    results['failed'] += 1
                    results['errors'].append(row_error)
                    continue
                
                # Get fees data
                total_fees_raw = row.get('Total_Fees')
                fees_paid_raw = row.get('Fees_Paid', 0)

                if pd.isna(total_fees_raw):
                    results['failed'] += 1
                    results['errors'].append(f"Row {index+2}: Total_Fees is required")
                    continue

                try:
                    total_fees = float(total_fees_raw)
                    fees_paid = 0.0 if pd.isna(fees_paid_raw) else float(fees_paid_raw)
                except (TypeError, ValueError):
                    results['failed'] += 1
                    results['errors'].append(f"Row {index+2}: Total_Fees and Fees_Paid must be valid numbers")
                    continue

                if total_fees < 0 or fees_paid < 0:
                    results['failed'] += 1
                    results['errors'].append(f"Row {index+2}: Fees cannot be negative")
                    continue
                
                # Validate fees
                if fees_paid > total_fees:
                    results['failed'] += 1
                    results['errors'].append(f"Row {index+2}: Fees paid cannot exceed total fees")
                    continue
                
                # Username is PRN, Password is Phone Number
                username = prn
                password = phone_number  # Password is the phone number
                
                # Create user
                user = User(
                    username=username,
                    password=generate_password_hash(password),
                    email=student_email,
                    phone=phone_number,
                    role='student',
                    consultancy_id=consultancy.id
                )
                db.session.add(user)
                db.session.flush()
                
                # Create student
                student = Student(
                    user_id=user.id,
                    consultancy_id=consultancy.id,
                    prn=prn,
                    full_name=student_name,
                    branch=student_branch,
                    email=student_email,
                    phone=phone_number,
                    total_fees=0,
                    fees_paid=0
                )
                db.session.add(student)
                db.session.flush()

                student.semester_fees.append(StudentSemesterFee(
                    semester=semester,
                    total_fees=total_fees,
                    fees_paid=fees_paid
                ))
                student.recalculate_fee_totals()

                seen_emails.add(student_email)
                seen_phones.add(phone_number)
                
                # Store credentials for display
                results['credentials'].append({
                    'prn': prn,
                    'name': student_name,
                    'username': username,
                    'password': password,  # Show phone number as password
                    'email': student_email,
                    'phone': phone_number,
                    'semester': semester
                })
                
                results['success'] += 1
                
            except Exception as e:
                results['failed'] += 1
                results['errors'].append(f"Row {index+2}: {str(e)}")
                db.session.rollback()
                continue
        
        db.session.commit()
        return True, results
        
    except Exception as e:
        db.session.rollback()
        return False, f"Error processing Excel file: {str(e)}"
    
def export_students_to_excel(students):
    """Export students data to Excel"""
    data = []
    for student in students:
        data.append({
            'PRN': student.prn,
            'Name': student.full_name,
            'Branch': student.branch,
            'Email': student.email,
            'Phone': student.phone,
            'Hostel_Code': f"{student.consultancy.hostel_code} - {student.consultancy.hostel_name}",
            'Total Fees': student.total_fees,
            'Fees Paid': student.fees_paid,
            'Fees Pending': student.fees_pending
        })
    
    df = pd.DataFrame(data)
    return df

def export_transactions_to_excel(transactions):
    """Export transactions to Excel"""
    data = []
    for txn in transactions:
        data.append({
            'Transaction ID': txn.transaction_id,
            'Student Name': txn.student.full_name,
            'PRN': txn.student.prn,
            'Branch': txn.student.branch,
            'Amount': txn.amount,
            'Payment Date': txn.payment_date.strftime('%Y-%m-%d %H:%M:%S'),
            'Status': txn.status
        })
    
    df = pd.DataFrame(data)
    # Sort alphabetically by student name
    df = df.sort_values('Student Name')
    return df
